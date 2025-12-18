"""
Script to create default Excel templates for inventory audit reports
"""

from pathlib import Path
from config.app_config import TEMPLATES_DIR
import logging

logger = logging.getLogger(__name__)

# Try to import openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available, cannot create Excel templates")


def _style_header(cell):
    """Style header cell"""
    if OPENPYXL_AVAILABLE:
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )


def create_default_templates():
    """Create default Excel templates"""
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl not available, cannot create templates")
        return
    
    excel_dir = TEMPLATES_DIR / "excel_templates"
    excel_dir.mkdir(parents=True, exist_ok=True)
    
    # Template 1: Inventory Overview
    wb = Workbook()
    ws = wb.active
    ws.title = "Készlet Áttekintés"
    
    ws['A1'] = "Készlet Áttekintés"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = "Időszak: {period}"
    ws['A3'] = "Dátum: {date}"
    
    row = 5
    headers = ["Mutató", "Érték"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        _style_header(cell)
    
    # Sample data rows
    sample_data = [
        ("Teljes készlet érték", "{total_stock_value}"),
        ("Teljes készlet mennyiség", "{total_stock_quantity}"),
        ("Aktív alkatrészek", "{active_parts}"),
        ("Alacsony készletű alkatrészek", "{low_stock_parts}"),
        ("Készleten kívüli alkatrészek", "{out_of_stock_parts}"),
    ]
    
    for data_row in sample_data:
        row += 1
        ws.cell(row, 1, data_row[0])
        ws.cell(row, 2, data_row[1])
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    template_path = excel_dir / "inventory_overview_template.xlsx"
    wb.save(str(template_path))
    logger.info(f"Created template: {template_path}")
    
    # Template 2: Usage Report
    wb = Workbook()
    ws = wb.active
    ws.title = "Felhasználás Kimutatás"
    
    ws['A1'] = "Felhasználás Kimutatás"
    ws['A1'].font = Font(bold=True, size=16)
    
    row = 3
    headers = ["Alkatrész neve", "SKU", "Mértékegység", "Felhasznált mennyiség", "Költség", "Trend", "Átlagos havi felhasználás"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        _style_header(cell)
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    template_path = excel_dir / "usage_report_template.xlsx"
    wb.save(str(template_path))
    logger.info(f"Created template: {template_path}")
    
    # Template 3: Stock Quantity Report
    wb = Workbook()
    ws = wb.active
    ws.title = "Készlet Mennyiség"
    
    ws['A1'] = "Készlet Mennyiség Kimutatás"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = "Dátum: {date}"
    
    row = 4
    headers = [
        "Alkatrész neve", "SKU", "Kategória", "Mértékegység",
        "Jelenlegi készlet", "Biztonsági készlet", "Újrarendelési mennyiség",
        "Beszerzési ár", "FIFO egységár", "Készlet érték",
        "Szállító", "Státusz", "Utolsó beérkezés", "Utolsó kiadás"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        _style_header(cell)
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    template_path = excel_dir / "stock_quantity_template.xlsx"
    wb.save(str(template_path))
    logger.info(f"Created template: {template_path}")
    
    logger.info(f"Default Excel templates created in {excel_dir}")


if __name__ == "__main__":
    create_default_templates()





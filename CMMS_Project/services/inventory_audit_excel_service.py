"""
Excel export service for inventory audit reports
"""

from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List
import logging

logger = logging.getLogger(__name__)

# Try to import openpyxl
OPENPYXL_AVAILABLE = False
Workbook = None
BarChart = None
LineChart = None
PieChart = None
try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError as e:
    logger.error(f"openpyxl ImportError: {e}")
    OPENPYXL_AVAILABLE = False

from services.inventory_audit_service import (
    get_inventory_overview,
    get_usage_report,
    get_value_change_report,
    get_stock_quantity_report,
    get_stock_change_report,
    get_machine_usage_trend,
    get_maintenance_trend_report,
)
from localization.translator import translator
from utils.currency import format_price

logger = logging.getLogger(__name__)


def _format_currency(value: float) -> str:
    """Format value as currency"""
    return f"{value:,.2f} €"


def _style_header(cell):
    """Style header cell"""
    if OPENPYXL_AVAILABLE:
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )


def _style_data_cell(cell):
    """Style data cell"""
    if OPENPYXL_AVAILABLE:
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )


def _style_number_cell(cell):
    """Style number cell"""
    if OPENPYXL_AVAILABLE:
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )


def export_inventory_overview_to_excel(
    period: str = "monthly",
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    machine_id: Optional[int] = None,
    output_path: Optional[Path] = None
) -> Path:
    """Export inventory overview to Excel"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")
    
    if output_path is None:
        output_dir = Path("generated_reports")
        output_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"inventory_overview_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Készlet Áttekintés"
    
    # Get data
    data = get_inventory_overview(period, start_date, end_date, machine_id)
    
    # Header
    ws['A1'] = "Készlet Áttekintés"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Időszak: {period}"
    ws['A3'] = f"Dátum: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Summary metrics
    row = 5
    ws[f'A{row}'] = "Mutató"
    ws[f'B{row}'] = "Érték"
    _style_header(ws[f'A{row}'])
    _style_header(ws[f'B{row}'])
    
    row += 1
    ws[f'A{row}'] = "Teljes készlet érték"
    ws[f'B{row}'] = _format_currency(data.get("total_stock_value", 0.0))
    _style_data_cell(ws[f'A{row}'])
    _style_number_cell(ws[f'B{row}'])
    
    row += 1
    ws[f'A{row}'] = "Teljes készlet mennyiség"
    ws[f'B{row}'] = data.get("total_stock_quantity", 0)
    _style_data_cell(ws[f'A{row}'])
    _style_number_cell(ws[f'B{row}'])
    
    row += 1
    ws[f'A{row}'] = "Aktív alkatrészek"
    ws[f'B{row}'] = data.get("active_parts", 0)
    _style_data_cell(ws[f'A{row}'])
    _style_number_cell(ws[f'B{row}'])
    
    row += 1
    ws[f'A{row}'] = "Alacsony készletű alkatrészek"
    ws[f'B{row}'] = data.get("low_stock_parts", 0)
    _style_data_cell(ws[f'A{row}'])
    _style_number_cell(ws[f'B{row}'])
    
    row += 1
    ws[f'A{row}'] = "Készleten kívüli alkatrészek"
    ws[f'B{row}'] = data.get("out_of_stock_parts", 0)
    _style_data_cell(ws[f'A{row}'])
    _style_number_cell(ws[f'B{row}'])
    
    # Top used parts
    row += 2
    ws[f'A{row}'] = "Legtöbbet felhasznált alkatrészek"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    ws[f'A{row}'] = "Alkatrész neve"
    ws[f'B{row}'] = "SKU"
    ws[f'C{row}'] = "Felhasznált mennyiség"
    _style_header(ws[f'A{row}'])
    _style_header(ws[f'B{row}'])
    _style_header(ws[f'C{row}'])
    
    for part in data.get("top_used_parts", [])[:10]:
        row += 1
        ws[f'A{row}'] = part.get("name", "")
        ws[f'B{row}'] = part.get("sku", "")
        ws[f'C{row}'] = part.get("total_used", 0)
        _style_data_cell(ws[f'A{row}'])
        _style_data_cell(ws[f'B{row}'])
        _style_number_cell(ws[f'C{row}'])
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(f"Inventory overview exported to {output_path}")
    return output_path


def export_usage_report_to_excel(
    period: str = "monthly",
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    machine_id: Optional[int] = None,
    part_id: Optional[int] = None,
    output_path: Optional[Path] = None
) -> Path:
    """Export usage report to Excel"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")
    
    if output_path is None:
        output_dir = Path("generated_reports")
        output_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"usage_report_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Felhasználás Kimutatás"
    
    data = get_usage_report(period, start_date, end_date, machine_id, part_id)
    
    # Header
    ws['A1'] = "Felhasználás Kimutatás"
    ws['A1'].font = Font(bold=True, size=16)
    
    # Table headers
    row = 3
    headers = ["Alkatrész neve", "SKU", "Mértékegység", "Felhasznált mennyiség", "Költség", "Trend", "Átlagos havi felhasználás"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        _style_header(cell)
    
    # Data rows
    for usage_item in data.get("usage_data", []):
        row += 1
        ws.cell(row, 1, usage_item.get("part_name", ""))
        ws.cell(row, 2, usage_item.get("sku", ""))
        ws.cell(row, 3, usage_item.get("unit", ""))
        ws.cell(row, 4, usage_item.get("total_used", 0))
        ws.cell(row, 5, _format_currency(usage_item.get("total_cost", 0.0)))
        ws.cell(row, 6, usage_item.get("trend", ""))
        ws.cell(row, 7, usage_item.get("avg_monthly_usage", 0.0))
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            if col in [4, 5, 7]:
                _style_number_cell(cell)
            else:
                _style_data_cell(cell)
    
    # Summary row
    row += 2
    ws.cell(row, 1, "Összesen:")
    ws.cell(row, 4, data.get("total_quantity_used", 0))
    ws.cell(row, 5, _format_currency(data.get("total_cost", 0.0)))
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 4).font = Font(bold=True)
    ws.cell(row, 5).font = Font(bold=True)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(f"Usage report exported to {output_path}")
    return output_path


def export_value_change_report_to_excel(
    period: str = "monthly",
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    output_path: Optional[Path] = None
) -> Path:
    """Export value change report to Excel"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")
    
    if output_path is None:
        output_dir = Path("generated_reports")
        output_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"value_change_report_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Értékváltozás"
    
    data = get_value_change_report(period, start_date, end_date)
    
    # Header
    ws['A1'] = "Értékváltozás Kimutatás"
    ws['A1'].font = Font(bold=True, size=16)
    
    # Table headers
    row = 3
    headers = ["Alkatrész neve", "SKU", "Kezdeti érték", "Végkészlet érték", "Értékváltozás", "Változás %", "Készletforgalmi arány"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        _style_header(cell)
    
    # Data rows
    for part_data in data.get("parts", []):
        row += 1
        ws.cell(row, 1, part_data.get("part_name", ""))
        ws.cell(row, 2, part_data.get("sku", ""))
        ws.cell(row, 3, _format_currency(part_data.get("initial_value", 0.0)))
        ws.cell(row, 4, _format_currency(part_data.get("final_value", 0.0)))
        ws.cell(row, 5, _format_currency(part_data.get("value_change", 0.0)))
        ws.cell(row, 6, f"{part_data.get('value_change_percent', 0.0):.2f}%")
        ws.cell(row, 7, f"{part_data.get('turnover_rate', 0.0):.2f}")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            if col in [3, 4, 5]:
                _style_number_cell(cell)
            else:
                _style_data_cell(cell)
    
    # Summary
    summary = data.get("summary", {})
    row += 2
    ws.cell(row, 1, "Összesítő:")
    ws.cell(row, 3, _format_currency(summary.get("total_initial_value", 0.0)))
    ws.cell(row, 4, _format_currency(summary.get("total_final_value", 0.0)))
    ws.cell(row, 5, _format_currency(summary.get("total_value_change", 0.0)))
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(f"Value change report exported to {output_path}")
    return output_path


def export_stock_quantity_report_to_excel(
    machine_id: Optional[int] = None,
    include_zero: bool = False,
    output_path: Optional[Path] = None
) -> Path:
    """Export stock quantity report to Excel"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")
    
    if output_path is None:
        output_dir = Path("generated_reports")
        output_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"stock_quantity_report_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Készlet Mennyiség"
    
    data = get_stock_quantity_report(machine_id, include_zero)
    
    # Header
    ws['A1'] = "Készlet Mennyiség Kimutatás"
    ws['A1'].font = Font(bold=True, size=16)
    
    # Table headers
    row = 3
    headers = ["Alkatrész neve", "SKU", "Kategória", "Jelenlegi készlet", "Biztonsági készlet", "Készlet érték", "Státusz", "Szállító", "Utolsó beérkezés", "Utolsó kiadás"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        _style_header(cell)
    
    # Data rows
    for part_data in data:
        row += 1
        ws.cell(row, 1, part_data.get("part_name", ""))
        ws.cell(row, 2, part_data.get("sku", ""))
        ws.cell(row, 3, part_data.get("category", ""))
        ws.cell(row, 4, part_data.get("current_quantity", 0))
        ws.cell(row, 5, part_data.get("safety_stock", 0))
        ws.cell(row, 6, _format_currency(part_data.get("stock_value", 0.0)))
        ws.cell(row, 7, part_data.get("status", ""))
        ws.cell(row, 8, part_data.get("supplier_name", "") or "")
        ws.cell(row, 9, part_data.get("last_receipt_date").strftime("%Y-%m-%d") if part_data.get("last_receipt_date") else "")
        ws.cell(row, 10, part_data.get("last_issue_date").strftime("%Y-%m-%d") if part_data.get("last_issue_date") else "")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            if col in [4, 5, 6]:
                _style_number_cell(cell)
            else:
                _style_data_cell(cell)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(f"Stock quantity report exported to {output_path}")
    return output_path


def export_stock_change_report_to_excel(
    period: str = "monthly",
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    part_id: Optional[int] = None,
    output_path: Optional[Path] = None
) -> Path:
    """Export stock change report to Excel"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")
    
    if output_path is None:
        output_dir = Path("generated_reports")
        output_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"stock_change_report_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Készlet Változás"
    
    data = get_stock_change_report(period, start_date, end_date, part_id)
    
    # Header
    ws['A1'] = "Készlet Változás Kimutatás"
    ws['A1'].font = Font(bold=True, size=16)
    
    # Table headers
    row = 3
    headers = ["Alkatrész neve", "SKU", "Kezdeti készlet", "Végkészlet", "Változás", "Változás %", "Beérkezések", "Kiadások", "Átlagos készlet", "Forgalmi sebesség (nap)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        _style_header(cell)
    
    # Data rows
    for part_data in data.get("parts", []):
        row += 1
        ws.cell(row, 1, part_data.get("part_name", ""))
        ws.cell(row, 2, part_data.get("sku", ""))
        ws.cell(row, 3, part_data.get("initial_quantity", 0))
        ws.cell(row, 4, part_data.get("final_quantity", 0))
        ws.cell(row, 5, part_data.get("quantity_change", 0))
        ws.cell(row, 6, f"{part_data.get('quantity_change_percent', 0.0):.2f}%")
        ws.cell(row, 7, part_data.get("received_quantity", 0))
        ws.cell(row, 8, part_data.get("issued_quantity", 0))
        ws.cell(row, 9, part_data.get("avg_quantity", 0.0))
        ws.cell(row, 10, f"{part_data.get('turnover_days', 0.0):.1f}")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            if col in [3, 4, 5, 7, 8, 9, 10]:
                _style_number_cell(cell)
            else:
                _style_data_cell(cell)
    
    # Summary
    summary = data.get("summary", {})
    row += 2
    ws.cell(row, 1, "Összesítő:")
    ws.cell(row, 3, summary.get("total_initial_quantity", 0))
    ws.cell(row, 4, summary.get("total_final_quantity", 0))
    ws.cell(row, 7, summary.get("total_received", 0))
    ws.cell(row, 8, summary.get("total_issued", 0))
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(f"Stock change report exported to {output_path}")
    return output_path


def export_machine_usage_trend_to_excel(
    machine_id: int,
    period: str = "monthly",
    breakdown: str = "monthly",
    output_path: Optional[Path] = None
) -> Path:
    """Export machine usage trend to Excel"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")
    
    if output_path is None:
        output_dir = Path("generated_reports")
        output_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"machine_usage_trend_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Eszköz Felhasználási Trend"
    
    data = get_machine_usage_trend(machine_id, period, breakdown)
    
    # Header
    ws['A1'] = f"Eszköz Felhasználási Trend - {data.get('machine_name', '')}"
    ws['A1'].font = Font(bold=True, size=16)
    
    # Breakdown data
    row = 3
    headers = ["Időszak", "Felhasznált mennyiség", "Költség", "Munkalapok száma"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        _style_header(cell)
    
    for breakdown_item in data.get("breakdown_data", []):
        row += 1
        ws.cell(row, 1, breakdown_item.get("period", ""))
        ws.cell(row, 2, breakdown_item.get("quantity_used", 0))
        ws.cell(row, 3, _format_currency(breakdown_item.get("cost", 0.0)))
        ws.cell(row, 4, breakdown_item.get("worksheet_count", 0))
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            if col in [2, 3, 4]:
                _style_number_cell(cell)
            else:
                _style_data_cell(cell)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(f"Machine usage trend exported to {output_path}")
    return output_path


def export_maintenance_trend_to_excel(
    machine_id: Optional[int] = None,
    period: str = "monthly",
    breakdown: str = "monthly",
    output_path: Optional[Path] = None
) -> Path:
    """Export maintenance trend to Excel"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")
    
    if output_path is None:
        output_dir = Path("generated_reports")
        output_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"maintenance_trend_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Karbantartási Trend"
    
    data = get_maintenance_trend_report(machine_id, period, breakdown)
    
    # Header
    ws['A1'] = "Karbantartási Trend Kimutatás"
    ws['A1'].font = Font(bold=True, size=16)
    
    # Breakdown data
    row = 3
    headers = ["Időszak", "Karbantartások száma", "Preventív", "Korrekciós", "Költség", "Leállási idő (óra)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        _style_header(cell)
    
    for breakdown_item in data.get("breakdown_data", []):
        row += 1
        ws.cell(row, 1, breakdown_item.get("period", ""))
        ws.cell(row, 2, breakdown_item.get("maintenance_count", 0))
        ws.cell(row, 3, breakdown_item.get("preventive_count", 0))
        ws.cell(row, 4, breakdown_item.get("corrective_count", 0))
        ws.cell(row, 5, _format_currency(breakdown_item.get("maintenance_cost", 0.0)))
        ws.cell(row, 6, breakdown_item.get("downtime_hours", 0.0))
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            if col in [2, 3, 4, 5, 6]:
                _style_number_cell(cell)
            else:
                _style_data_cell(cell)
    
    # Summary
    summary = data.get("summary", {})
    row += 2
    ws.cell(row, 1, "Összesítő:")
    ws.cell(row, 2, summary.get("total_maintenances", 0))
    ws.cell(row, 3, summary.get("preventive_count", 0))
    ws.cell(row, 4, summary.get("corrective_count", 0))
    ws.cell(row, 5, _format_currency(summary.get("total_cost", 0.0)))
    ws.cell(row, 6, summary.get("total_downtime_hours", 0.0))
    
    row += 1
    ws.cell(row, 1, "MTBF (óra):")
    ws.cell(row, 2, summary.get("mtbf_hours", 0.0))
    
    row += 1
    ws.cell(row, 1, "MTTR (óra):")
    ws.cell(row, 2, summary.get("mttr_hours", 0.0))
    
    row += 1
    ws.cell(row, 1, "Rendelkezésre állás (%):")
    ws.cell(row, 2, f"{summary.get('availability_percent', 0.0):.2f}%")
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(f"Maintenance trend exported to {output_path}")
    return output_path


def export_full_inventory_to_excel(
    output_path: Optional[Path] = None
) -> Path:
    """Export full inventory with all information to Excel"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")
    
    if output_path is None:
        output_dir = Path("generated_reports")
        output_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"full_inventory_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Teljes Készlet"
    
    data = get_stock_quantity_report(include_zero=True)
    
    # Header
    ws['A1'] = "Teljes Készlet Kimutatás"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Dátum: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Table headers
    row = 4
    headers = [
        "Alkatrész neve", "SKU", "Kategória", "Mértékegység",
        "Jelenlegi készlet", "Biztonsági készlet", "Újrarendelési mennyiség",
        "Beszerzési ár", "FIFO egységár", "Készlet érték",
        "Szállító", "Státusz", "Utolsó beérkezés", "Utolsó kiadás"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        _style_header(cell)
    
    # Data rows
    for part_data in data:
        row += 1
        ws.cell(row, 1, part_data.get("part_name", ""))
        ws.cell(row, 2, part_data.get("sku", ""))
        ws.cell(row, 3, part_data.get("category", ""))
        ws.cell(row, 4, part_data.get("unit", ""))
        ws.cell(row, 5, part_data.get("current_quantity", 0))
        ws.cell(row, 6, part_data.get("safety_stock", 0))
        ws.cell(row, 7, part_data.get("reorder_quantity", 0))
        ws.cell(row, 8, _format_currency(part_data.get("buy_price", 0.0)))
        ws.cell(row, 9, _format_currency(part_data.get("fifo_unit_cost", 0.0)))
        ws.cell(row, 10, _format_currency(part_data.get("stock_value", 0.0)))
        ws.cell(row, 11, part_data.get("supplier_name", "") or "")
        ws.cell(row, 12, part_data.get("status", ""))
        ws.cell(row, 13, part_data.get("last_receipt_date").strftime("%Y-%m-%d") if part_data.get("last_receipt_date") else "")
        ws.cell(row, 14, part_data.get("last_issue_date").strftime("%Y-%m-%d") if part_data.get("last_issue_date") else "")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            if col in [5, 6, 7, 8, 9, 10]:
                _style_number_cell(cell)
            else:
                _style_data_cell(cell)
    
    # Summary row
    row += 2
    total_value = sum(p.get("stock_value", 0.0) for p in data)
    total_quantity = sum(p.get("current_quantity", 0) for p in data)
    ws.cell(row, 1, "Összesen:")
    ws.cell(row, 5, total_quantity)
    ws.cell(row, 10, _format_currency(total_value))
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 5).font = Font(bold=True)
    ws.cell(row, 10).font = Font(bold=True)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(f"Full inventory exported to {output_path}")
    return output_path


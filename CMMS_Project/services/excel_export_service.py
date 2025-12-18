"""
Excel export service for reports with charts
"""

from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List
import logging

logger = logging.getLogger(__name__)

# Try to import openpyxl with detailed error logging
OPENPYXL_AVAILABLE = False
Workbook = None
BarChart = None
LineChart = None
PieChart = None
try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
    logger.info("openpyxl successfully imported")
except ImportError as e:
    OPENPYXL_AVAILABLE = False
    Workbook = None
    BarChart = None
    LineChart = None
    PieChart = None
    logger.error(f"openpyxl ImportError: {e}")
    import sys
    logger.error(f"Python executable: {sys.executable}")
    logger.error(f"Python path: {sys.path[:3]}")
except Exception as e:
    OPENPYXL_AVAILABLE = False
    Workbook = None
    BarChart = None
    LineChart = None
    PieChart = None
    logger.error(f"openpyxl import failed with exception: {type(e).__name__}: {e}")
    import sys
    logger.error(f"Python executable: {sys.executable}")
    logger.error(f"Python path: {sys.path[:3]}")

from services.reports_service import get_all_statistics, get_period_comparison
from localization.translator import translator
import csv

logger = logging.getLogger(__name__)


def _format_currency(value: float) -> str:
    """Format value as currency in EUR"""
    return f"{value:,.2f} €"


def _format_hours(value: float) -> str:
    """Format value as hours"""
    return f"{value:.2f} óra"


def _style_header(cell):
    """Style header cell"""
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def _style_data_cell(cell):
    """Style data cell"""
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def _style_number_cell(cell):
    """Style number cell"""
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def export_reports_to_excel(
    periods: List[str] = ["day", "week", "month", "year"],
    user_id: Optional[int] = None,
    output_path: Optional[Path] = None
) -> Path:
    """Export reports to Excel with charts"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")
    
    # user_id is passed as parameter, no need to get from context here
    
    if output_path is None:
        output_dir = Path("generated_pdfs")
        output_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"reports_{timestamp}.xlsx"
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Get statistics for all periods
    all_stats = get_period_comparison(periods, user_id)
    
    # Create summary sheet
    _create_summary_sheet(wb, all_stats, periods, user_id)
    
    # Create cost sheet with chart
    _create_cost_sheet(wb, all_stats, periods, user_id)
    
    # Create time sheet with chart
    _create_time_sheet(wb, all_stats, periods, user_id)
    
    # Create tasks sheet with chart
    _create_tasks_sheet(wb, all_stats, periods, user_id)
    
    # Create detailed breakdown sheet
    _create_detailed_sheet(wb, all_stats, periods, user_id)
    
    # Ensure parent directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Save workbook
    wb.save(str(output_path))
    
    # Verify file was created
    if not output_path.exists():
        raise IOError(f"Failed to create Excel file at {output_path}")
    
    logger.info(f"Excel report exported to {output_path}")
    return output_path


def export_reports_to_csv(
    periods: List[str] = ["day", "week", "month", "year"],
    user_id: Optional[int] = None,
    output_path: Optional[Path] = None
) -> Path:
    """Export reports to CSV format"""
    if output_path is None:
        output_dir = Path("generated_reports")
        output_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"reports_{timestamp}.csv"
    
    # Get statistics for all periods
    all_stats = get_period_comparison(periods, user_id)
    
    # Ensure parent directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Write CSV file
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        
        # Header row
        writer.writerow([
            'Period',
            'Total Cost',
            'Worksheet Cost',
            'Service Cost',
            'Total Time (hours)',
            'Worksheet Downtime (hours)',
            'PM Duration (hours)',
            'Service Duration (hours)',
            'Total Tasks',
            'Worksheet Count',
            'PM Count',
            'Service Count'
        ])
        
        # Data rows
        for period in periods:
            period_stats = all_stats.get(period, {})
            cost_stats = period_stats.get('cost', {})
            time_stats = period_stats.get('time', {})
            task_stats = period_stats.get('tasks', {})
            
            writer.writerow([
                period,
                cost_stats.get('total_cost', 0),
                cost_stats.get('worksheet_cost', 0),
                cost_stats.get('service_cost', 0),
                time_stats.get('total_time_hours', 0),
                time_stats.get('worksheet_downtime_hours', 0),
                time_stats.get('pm_duration_hours', 0),
                time_stats.get('service_duration_hours', 0),
                task_stats.get('total_tasks', 0),
                task_stats.get('worksheet_count', 0),
                task_stats.get('pm_count', 0),
                task_stats.get('service_count', 0)
            ])
    
    logger.info(f"CSV report exported to {output_path}")
    return output_path


def _create_summary_sheet(wb: Workbook, all_stats: Dict, periods: List[str], user_id: Optional[int]):
    """Create summary sheet"""
    ws = wb.create_sheet("Összefoglaló")
    
    # Title
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = translator.get_text("reports.title") if hasattr(translator, 'get_text') else "Grafikonok és Kimutatások / Reports and Charts"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Date and user info
    ws['A2'] = f"Generálva / Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    if user_id:
        try:
            from services.context_service import get_app_context
            ctx = get_app_context()
            if ctx:
                ws['A3'] = f"Felhasználó / User: {ctx.username or 'N/A'}"
        except:
            pass
    
    row = 5
    
    # Summary table
    ws[f'A{row}'] = "Időszak / Period"
    ws[f'B{row}'] = "Össz Költség / Total Cost"
    ws[f'C{row}'] = "Össz Idő / Total Time"
    ws[f'D{row}'] = "Össz Feladat / Total Tasks"
    
    for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}']]:
        _style_header(cell)
    
    row += 1
    
    period_labels = {
        "day": "Nap / Day",
        "week": "Hét / Week",
        "month": "Hó / Month",
        "year": "Év / Year"
    }
    
    for period in periods:
        stats = all_stats.get(period, {})
        cost_stats = stats.get('cost', {})
        time_stats = stats.get('time', {})
        task_stats = stats.get('tasks', {})
        
        ws[f'A{row}'] = period_labels.get(period, period)
        ws[f'B{row}'] = _format_currency(cost_stats.get('total_cost', 0))
        ws[f'C{row}'] = _format_hours(time_stats.get('total_time_hours', 0))
        ws[f'D{row}'] = task_stats.get('total_tasks', 0)
        
        _style_data_cell(ws[f'A{row}'])
        _style_number_cell(ws[f'B{row}'])
        _style_number_cell(ws[f'C{row}'])
        _style_number_cell(ws[f'D{row}'])
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


def _create_cost_sheet(wb: Workbook, all_stats: Dict, periods: List[str], user_id: Optional[int]):
    """Create cost analysis sheet with chart"""
    ws = wb.create_sheet("Költség")
    
    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "Költség Elemzés / Cost Analysis"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    row = 3
    
    # Headers
    headers = ["Időszak / Period", "Munkalap Költség / Worksheet Cost", "Szerviz Költség / Service Cost", "Összesen / Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        _style_header(cell)
    
    row += 1
    
    period_labels = {
        "day": "Nap / Day",
        "week": "Hét / Week",
        "month": "Hó / Month",
        "year": "Év / Year"
    }
    
    for period in periods:
        stats = all_stats.get(period, {})
        cost_stats = stats.get('cost', {})
        
        ws.cell(row=row, column=1).value = period_labels.get(period, period)
        ws.cell(row=row, column=2).value = cost_stats.get('worksheet_cost', 0)
        ws.cell(row=row, column=3).value = cost_stats.get('service_cost', 0)
        ws.cell(row=row, column=4).value = cost_stats.get('total_cost', 0)
        
        _style_data_cell(ws.cell(row=row, column=1))
        _style_number_cell(ws.cell(row=row, column=2))
        _style_number_cell(ws.cell(row=row, column=3))
        _style_number_cell(ws.cell(row=row, column=4))
        
        row += 1
    
    # Create bar chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Költség Időszakonként / Cost by Period"
    chart.y_axis.title = "Költség (€) / Cost (€)"
    chart.x_axis.title = "Időszak / Period"
    
    data = Reference(ws, min_col=2, min_row=3, max_col=4, max_row=row-1)
    cats = Reference(ws, min_col=1, min_row=4, max_row=row-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 15
    
    ws.add_chart(chart, "F3")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25


def _create_time_sheet(wb: Workbook, all_stats: Dict, periods: List[str], user_id: Optional[int]):
    """Create time analysis sheet with chart"""
    ws = wb.create_sheet("Idő")
    
    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "Idő Elemzés / Time Analysis"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    row = 3
    
    # Headers
    headers = ["Időszak / Period", "Munkalap Leállás / Worksheet Downtime", "PM Idő / PM Duration", "Szerviz Idő / Service Duration", "Összesen / Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        _style_header(cell)
    
    row += 1
    
    period_labels = {
        "day": "Nap / Day",
        "week": "Hét / Week",
        "month": "Hó / Month",
        "year": "Év / Year"
    }
    
    for period in periods:
        stats = all_stats.get(period, {})
        time_stats = stats.get('time', {})
        
        ws.cell(row=row, column=1).value = period_labels.get(period, period)
        ws.cell(row=row, column=2).value = time_stats.get('worksheet_downtime_hours', 0)
        ws.cell(row=row, column=3).value = time_stats.get('pm_duration_hours', 0)
        ws.cell(row=row, column=4).value = time_stats.get('service_duration_hours', 0)
        ws.cell(row=row, column=5).value = time_stats.get('total_time_hours', 0)
        
        _style_data_cell(ws.cell(row=row, column=1))
        _style_number_cell(ws.cell(row=row, column=2))
        _style_number_cell(ws.cell(row=row, column=3))
        _style_number_cell(ws.cell(row=row, column=4))
        _style_number_cell(ws.cell(row=row, column=5))
        
        row += 1
    
    # Create line chart
    chart = LineChart()
    chart.title = "Idő Időszakonként / Time by Period"
    chart.y_axis.title = "Idő (óra) / Time (hours)"
    chart.x_axis.title = "Időszak / Period"
    chart.style = 10
    
    # Data reference: columns 2-5 (downtime, pm_time, service_time, total), rows 3 (header) to row-1 (last data)
    data = Reference(ws, min_col=2, min_row=3, max_col=5, max_row=row-1)
    # Categories reference: column 1 (period), rows 4 to row-1 (data rows, skip header)
    cats = Reference(ws, min_col=1, min_row=4, max_row=row-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 18
    
    ws.add_chart(chart, "F3")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 25


def _create_tasks_sheet(wb: Workbook, all_stats: Dict, periods: List[str], user_id: Optional[int]):
    """Create tasks analysis sheet with chart"""
    ws = wb.create_sheet("Feladatok")
    
    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "Feladat Elemzés / Task Analysis"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    row = 3
    
    # Headers
    headers = ["Időszak / Period", "Munkalapok / Worksheets", "PM Feladatok / PM Tasks", "Összesen / Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        _style_header(cell)
    
    row += 1
    
    period_labels = {
        "day": "Nap / Day",
        "week": "Hét / Week",
        "month": "Hó / Month",
        "year": "Év / Year"
    }
    
    for period in periods:
        stats = all_stats.get(period, {})
        task_stats = stats.get('tasks', {})
        
        ws.cell(row=row, column=1).value = period_labels.get(period, period)
        ws.cell(row=row, column=2).value = task_stats.get('worksheet_count', 0)
        ws.cell(row=row, column=3).value = task_stats.get('pm_count', 0)
        ws.cell(row=row, column=4).value = task_stats.get('total_tasks', 0)
        
        _style_data_cell(ws.cell(row=row, column=1))
        _style_number_cell(ws.cell(row=row, column=2))
        _style_number_cell(ws.cell(row=row, column=3))
        _style_number_cell(ws.cell(row=row, column=4))
        
        row += 1
    
    # Bar chart for all periods (first chart)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Feladatok Időszakonként / Tasks by Period"
    chart.y_axis.title = "Feladatok száma / Number of Tasks"
    chart.x_axis.title = "Időszak / Period"
    
    data = Reference(ws, min_col=2, min_row=3, max_col=4, max_row=row-1)
    cats = Reference(ws, min_col=1, min_row=4, max_row=row-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 15
    
    ws.add_chart(chart, "F3")
    
    # Create pie chart for last period (usually year)
    if periods:
        last_period = periods[-1]
        stats = all_stats.get(last_period, {})
        task_stats = stats.get('tasks', {})
        
        # Create pie chart data in a separate area
        pie_row = row + 2
        ws.cell(row=pie_row, column=1).value = "Feladat Típus / Task Type"
        ws.cell(row=pie_row, column=2).value = "Mennyiség / Count"
        _style_header(ws.cell(row=pie_row, column=1))
        _style_header(ws.cell(row=pie_row, column=2))
        
        pie_row += 1
        ws.cell(row=pie_row, column=1).value = "Munkalapok / Worksheets"
        ws.cell(row=pie_row, column=2).value = task_stats.get('worksheet_count', 0)
        _style_data_cell(ws.cell(row=pie_row, column=1))
        _style_number_cell(ws.cell(row=pie_row, column=2))
        
        pie_row += 1
        ws.cell(row=pie_row, column=1).value = "PM Feladatok / PM Tasks"
        ws.cell(row=pie_row, column=2).value = task_stats.get('pm_count', 0)
        _style_data_cell(ws.cell(row=pie_row, column=1))
        _style_number_cell(ws.cell(row=pie_row, column=2))
        
        # Create pie chart - fix data reference
        pie_chart = PieChart()
        pie_chart.title = f"Feladat Eloszlás ({last_period}) / Task Distribution ({last_period})"
        # Data reference: column 2, from header row to last data row
        data = Reference(ws, min_col=2, min_row=pie_row-1, max_row=pie_row)
        # Categories reference: column 1, from header row to last data row
        cats = Reference(ws, min_col=1, min_row=pie_row-1, max_row=pie_row)
        pie_chart.add_data(data, titles_from_data=False)
        pie_chart.set_categories(cats)
        pie_chart.height = 10
        pie_chart.width = 10
        # Place pie chart next to bar chart
        ws.add_chart(pie_chart, "F20")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 25


def _create_detailed_sheet(wb: Workbook, all_stats: Dict, periods: List[str], user_id: Optional[int]):
    """Create detailed breakdown sheet"""
    ws = wb.create_sheet("Részletes")
    
    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "Részletes Kimutatás / Detailed Report"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    row = 3
    
    # Detailed headers
    headers = [
        "Időszak / Period",
        "Munkalap Költség / Worksheet Cost",
        "Szerviz Költség / Service Cost",
        "Össz Költség / Total Cost",
        "Leállás (óra) / Downtime (h)",
        "PM Idő (óra) / PM Time (h)",
        "Munkalapok / Worksheets",
        "PM Feladatok / PM Tasks"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        _style_header(cell)
    
    row += 1
    
    period_labels = {
        "day": "Nap / Day",
        "week": "Hét / Week",
        "month": "Hó / Month",
        "year": "Év / Year"
    }
    
    for period in periods:
        stats = all_stats.get(period, {})
        cost_stats = stats.get('cost', {})
        time_stats = stats.get('time', {})
        task_stats = stats.get('tasks', {})
        
        ws.cell(row=row, column=1).value = period_labels.get(period, period)
        ws.cell(row=row, column=2).value = cost_stats.get('worksheet_cost', 0)
        ws.cell(row=row, column=3).value = cost_stats.get('service_cost', 0)
        ws.cell(row=row, column=4).value = cost_stats.get('total_cost', 0)
        ws.cell(row=row, column=5).value = time_stats.get('worksheet_downtime_hours', 0)
        ws.cell(row=row, column=6).value = time_stats.get('pm_duration_hours', 0)
        ws.cell(row=row, column=7).value = task_stats.get('worksheet_count', 0)
        ws.cell(row=row, column=8).value = task_stats.get('pm_count', 0)
        
        _style_data_cell(ws.cell(row=row, column=1))
        for col in range(2, 9):
            _style_number_cell(ws.cell(row=row, column=col))
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 25

